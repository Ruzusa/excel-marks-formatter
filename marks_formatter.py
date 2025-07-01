from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Load workbook and sheet
wb = load_workbook("marks.xlsx")
ws = wb.active

# Add headers for Total and Grade
ws.cell(row=1, column=5, value="Total")
ws.cell(row=1, column=6, value="Grade")

# Bold the header row
for cell in ws[1]:
    cell.font = Font(bold=True)

# Loop through each student row
for row in range(2, ws.max_row + 1):
    try:
        math = int(ws.cell(row=row, column=2).value or 0)
        physics = int(ws.cell(row=row, column=3).value or 0)
        chemistry = int(ws.cell(row=row, column=4).value or 0)
    except ValueError:
        print(f"⚠️ Skipping row {row} due to invalid data.")
        continue

    total = math + physics + chemistry
    ws.cell(row=row, column=5, value=total)

    # Calculate Grade
    if total >= 240:
        grade = "A"
        fill = PatternFill(start_color="C6EFCE",
                           end_color="C6EFCE", fill_type="solid")
    elif total >= 180:
        grade = "B"
        fill = PatternFill(start_color="FFEB9C",
                           end_color="FFEB9C", fill_type="solid")
    else:
        grade = "C"
        fill = PatternFill(start_color="F4CCCC",
                           end_color="F4CCCC", fill_type="solid")

    grade_cell = ws.cell(row=row, column=6, value=grade)
    grade_cell.fill = fill


# Save the workbook
wb.save("formatted_marks.xlsx")

print("✅ Excel file formatted and saved as 'formatted_marks.xlsx'")
